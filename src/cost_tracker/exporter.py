from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from cost_tracker.config import Settings
from cost_tracker.db import get_assignees_with_cost, get_conn, get_issues_with_cost

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1A3A5C")


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _autofit(ws: Worksheet) -> None:
    for col in ws.columns:  # type: ignore[attr-defined]
        width = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column or 1)].width = min(width + 4, 60)


def export_excel(settings: Settings) -> Path:
    Path(settings.export_dir).mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H-%M")
    out_path = Path(settings.export_dir) / f"cost_{ts}.xlsx"

    with get_conn(settings.db_path) as conn:
        issues = get_issues_with_cost(conn)
        assignees = get_assignees_with_cost(conn)
        raw_worklogs = conn.execute(
            """SELECT w.worklog_id, w.issue_key, w.project_key, w.issue_summary,
                      w.author_display_name,
                      ROUND(w.time_spent_seconds / 3600.0, 2) AS hours,
                      ROUND(w.time_spent_seconds / 3600.0 * COALESCE(r.rate_eur, 0), 2) AS cost_eur,
                      w.started, w.synced_at
               FROM worklogs w
               LEFT JOIN hourly_rates r ON w.author_account_id = r.account_id
               ORDER BY w.started DESC"""
        ).fetchall()

    wb = openpyxl.Workbook()

    ws1: Worksheet = wb.active  # type: ignore[assignment]
    ws1.title = "By Issue"
    _write_header(ws1, ["Key", "Summary", "Project", "Hours", "Cost (€)"])
    for r in issues:
        ws1.append([r["issue_key"], r["issue_summary"], r["project_key"], r["hours"], r["cost_eur"]])
    _autofit(ws1)

    ws2: Worksheet = wb.create_sheet("By Assignee")
    _write_header(ws2, ["Assignee", "Issues", "Days", "Cost (€)", "Rate (€/h)"])
    for r in assignees:
        ws2.append([r["display_name"], r["issue_count"], r["man_days"], r["cost_eur"],
                    r["rate_eur"] if r["rate_eur"] is not None else "—"])
    _autofit(ws2)

    ws3: Worksheet = wb.create_sheet("Worklogs")
    _write_header(ws3, ["Worklog ID", "Issue", "Project", "Summary", "Author",
                        "Hours", "Cost (€)", "Started", "Synced At"])
    for r in raw_worklogs:
        ws3.append(list(r))
    _autofit(ws3)

    wb.save(out_path)
    return out_path
