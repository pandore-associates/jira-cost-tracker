from pathlib import Path

import openpyxl
import pytest

from cost_tracker.config import Settings
from cost_tracker.db import get_conn, init_db, set_rate, upsert_author, upsert_worklog
from cost_tracker.exporter import export_excel
from cost_tracker.jira_client import WorklogEntry


@pytest.fixture
def settings(env_defaults: None) -> Settings:
    return Settings()


@pytest.fixture(autouse=True)
def seeded_db(settings: Settings) -> None:
    init_db(settings.db_path)
    with get_conn(settings.db_path) as conn:
        upsert_author(conn, "acc1", "Alice", "2026-05-11T10:00:00")
        set_rate(conn, "acc1", 100.0, "2026-05-11T10:00:00")
        upsert_worklog(
            conn,
            WorklogEntry(
                worklog_id="wl1",
                issue_key="CSP-1",
                project_key="CSP",
                issue_summary="Test issue",
                author_account_id="acc1",
                author_display_name="Alice",
                time_spent_seconds=3600,
                started="2026-05-11T09:00:00.000+0000",
                assignee_account_id="acc1",
                assignee_display_name="Alice",
            ),
            "2026-05-11T10:00:00",
        )


def test_export_creates_xlsx_file(settings: Settings) -> None:
    path = export_excel(settings)
    assert path.exists()
    assert path.suffix == ".xlsx"


def test_export_has_three_sheets(settings: Settings) -> None:
    path = export_excel(settings)
    wb = openpyxl.load_workbook(path)
    assert set(wb.sheetnames) == {"By Issue", "By Assignee", "Worklogs"}


def test_export_by_issue_data(settings: Settings) -> None:
    path = export_excel(settings)
    ws = openpyxl.load_workbook(path)["By Issue"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0][0] == "CSP-1"   # Key
    assert rows[0][3] == 1.0       # Hours
    assert rows[0][4] == 100.0     # Cost


def test_export_by_assignee_data(settings: Settings) -> None:
    path = export_excel(settings)
    ws = openpyxl.load_workbook(path)["By Assignee"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0][0] == "Alice"
    assert rows[0][2] == 1       # 1 man-day (1 h ceiled)
    assert rows[0][3] == 800.0   # 1 day × €100/h × 8 h


def test_export_worklogs_sheet_has_raw_data(settings: Settings) -> None:
    path = export_excel(settings)
    ws = openpyxl.load_workbook(path)["Worklogs"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0][0] == "wl1"    # worklog_id
